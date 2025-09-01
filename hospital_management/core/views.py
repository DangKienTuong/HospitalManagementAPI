"""
Base views and mixins for the Hospital Management System.
Provides reusable view components following DRY principle.
"""

from rest_framework import viewsets, status, mixins
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import QuerySet
from django.core.cache import cache
from typing import Any, Dict, Optional, Type
import logging

from .pagination import CustomPageNumberPagination
from .exceptions import ResourceNotFoundException, ValidationException
from .services.base import BaseService
from .repositories.base import BaseRepository

logger = logging.getLogger(__name__)


class BaseViewSet(viewsets.ModelViewSet):
    """
    Base ViewSet with common functionality.
    All ViewSets should inherit from this for consistency.
    """
    
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPageNumberPagination
    
    # Service and repository for clean architecture
    service_class: Optional[Type[BaseService]] = None
    repository_class: Optional[Type[BaseRepository]] = None
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        # Initialize service and repository if provided
        if self.service_class:
            self.service = self.service_class()
        
        if self.repository_class:
            self.repository = self.repository_class()
    
    def get_queryset(self) -> QuerySet:
        """
        Get queryset with optimizations.
        Automatically applies select_related and prefetch_related.
        """
        queryset = super().get_queryset()
        
        # Apply select_related if defined
        if hasattr(self, 'select_related_fields'):
            queryset = queryset.select_related(*self.select_related_fields)
        
        # Apply prefetch_related if defined
        if hasattr(self, 'prefetch_related_fields'):
            queryset = queryset.prefetch_related(*self.prefetch_related_fields)
        
        # Filter soft deleted records by default
        if hasattr(queryset.model, 'is_deleted'):
            queryset = queryset.filter(is_deleted=False)
        
        return queryset
    
    def perform_create(self, serializer):
        """Add audit information on create."""
        kwargs = {}
        
        # Add current user to audit fields
        if hasattr(serializer.Meta.model, 'created_by'):
            kwargs['created_by'] = self.request.user
        
        serializer.save(**kwargs)
    
    def perform_update(self, serializer):
        """Add audit information on update."""
        kwargs = {}
        
        # Add current user to audit fields
        if hasattr(serializer.Meta.model, 'updated_by'):
            kwargs['updated_by'] = self.request.user
        
        serializer.save(**kwargs)
    
    def perform_destroy(self, instance):
        """Perform soft delete by default."""
        if hasattr(instance, 'delete'):
            # Use soft delete if available
            instance.delete(soft=True, user=self.request.user)
        else:
            super().perform_destroy(instance)
    
    def get_success_response(self, data: Any = None, message: str = "Success", 
                           status_code: int = status.HTTP_200_OK) -> Response:
        """
        Create standardized success response.
        
        Args:
            data: Response data
            message: Success message
            status_code: HTTP status code
            
        Returns:
            Response object
        """
        response_data = {
            'success': True,
            'message': message,
        }
        
        if data is not None:
            response_data['data'] = data
        
        return Response(response_data, status=status_code)
    
    def get_error_response(self, error: str, status_code: int = status.HTTP_400_BAD_REQUEST,
                          errors: Optional[Dict] = None) -> Response:
        """
        Create standardized error response.
        
        Args:
            error: Error message
            status_code: HTTP status code
            errors: Field-specific errors
            
        Returns:
            Response object
        """
        response_data = {
            'success': False,
            'error': error,
        }
        
        if errors:
            response_data['errors'] = errors
        
        return Response(response_data, status=status_code)


class CachedViewMixin:
    """
    Mixin for view-level caching.
    Caches GET requests for improved performance.
    """
    
    cache_timeout = 300  # 5 minutes default
    cache_key_prefix = 'view'
    
    def get_cache_key(self, request) -> str:
        """Generate cache key for request."""
        return f"{self.cache_key_prefix}:{request.path}:{request.query_params}"
    
    def list(self, request, *args, **kwargs):
        """List with caching."""
        # Try to get from cache
        cache_key = self.get_cache_key(request)
        cached_data = cache.get(cache_key)
        
        if cached_data is not None:
            logger.debug(f"Cache hit for {cache_key}")
            return Response(cached_data)
        
        # Get fresh data
        response = super().list(request, *args, **kwargs)
        
        # Cache successful responses
        if response.status_code == 200:
            cache.set(cache_key, response.data, self.cache_timeout)
        
        return response
    
    def retrieve(self, request, *args, **kwargs):
        """Retrieve with caching."""
        # Generate cache key with pk
        pk = kwargs.get('pk')
        cache_key = f"{self.cache_key_prefix}:{self.basename}:{pk}"
        
        # Try to get from cache
        cached_data = cache.get(cache_key)
        
        if cached_data is not None:
            logger.debug(f"Cache hit for {cache_key}")
            return Response(cached_data)
        
        # Get fresh data
        response = super().retrieve(request, *args, **kwargs)
        
        # Cache successful responses
        if response.status_code == 200:
            cache.set(cache_key, response.data, self.cache_timeout)
        
        return response
    
    def invalidate_cache(self, pk: Optional[Any] = None):
        """Invalidate cache for this view."""
        if pk:
            cache_key = f"{self.cache_key_prefix}:{self.basename}:{pk}"
            cache.delete(cache_key)
        else:
            # Clear all cache for this view
            pattern = f"{self.cache_key_prefix}:{self.basename}:*"
            cache.delete_pattern(pattern)


class BulkOperationMixin:
    """
    Mixin for bulk operations.
    Supports bulk create, update, and delete.
    """
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        Bulk create multiple records.
        
        Expected data format:
        {
            "items": [
                {...},
                {...}
            ]
        }
        """
        items = request.data.get('items', [])
        
        if not items:
            return self.get_error_response("No items provided")
        
        serializer = self.get_serializer(data=items, many=True)
        
        if serializer.is_valid():
            self.perform_bulk_create(serializer)
            return self.get_success_response(
                serializer.data,
                f"Created {len(items)} records",
                status.HTTP_201_CREATED
            )
        
        return self.get_error_response("Validation failed", errors=serializer.errors)
    
    @action(detail=False, methods=['patch'])
    def bulk_update(self, request):
        """
        Bulk update multiple records.
        
        Expected data format:
        {
            "items": [
                {"id": 1, "field": "value"},
                {"id": 2, "field": "value"}
            ]
        }
        """
        items = request.data.get('items', [])
        
        if not items:
            return self.get_error_response("No items provided")
        
        updated = []
        errors = []
        
        for item in items:
            pk = item.pop('id', None)
            if not pk:
                errors.append({"error": "ID required for update"})
                continue
            
            try:
                instance = self.get_queryset().get(pk=pk)
                serializer = self.get_serializer(instance, data=item, partial=True)
                
                if serializer.is_valid():
                    self.perform_update(serializer)
                    updated.append(serializer.data)
                else:
                    errors.append({"id": pk, "errors": serializer.errors})
            
            except self.queryset.model.DoesNotExist:
                errors.append({"id": pk, "error": "Not found"})
        
        return self.get_success_response({
            "updated": updated,
            "errors": errors
        }, f"Updated {len(updated)} records")
    
    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        """
        Bulk delete multiple records.
        
        Expected data format:
        {
            "ids": [1, 2, 3]
        }
        """
        ids = request.data.get('ids', [])
        
        if not ids:
            return self.get_error_response("No IDs provided")
        
        queryset = self.get_queryset().filter(pk__in=ids)
        count = queryset.count()
        
        for instance in queryset:
            self.perform_destroy(instance)
        
        return self.get_success_response(
            message=f"Deleted {count} records",
            status_code=status.HTTP_204_NO_CONTENT
        )
    
    def perform_bulk_create(self, serializer):
        """Perform bulk create with audit fields."""
        kwargs = {}
        
        if hasattr(serializer.child.Meta.model, 'created_by'):
            kwargs['created_by'] = self.request.user
        
        serializer.save(**kwargs)


class ExportMixin:
    """
    Mixin for data export functionality.
    Supports CSV, Excel, and PDF export.
    """
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export data in specified format.
        
        Query params:
        - format: csv, excel, pdf
        - fields: comma-separated list of fields to export
        """
        export_format = request.query_params.get('format', 'csv')
        fields = request.query_params.get('fields', '').split(',')
        
        queryset = self.filter_queryset(self.get_queryset())
        
        if export_format == 'csv':
            return self.export_csv(queryset, fields)
        elif export_format == 'excel':
            return self.export_excel(queryset, fields)
        elif export_format == 'pdf':
            return self.export_pdf(queryset, fields)
        else:
            return self.get_error_response(f"Unsupported format: {export_format}")
    
    def export_csv(self, queryset, fields):
        """Export data as CSV."""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.basename}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        if fields:
            writer.writerow(fields)
        else:
            fields = [f.name for f in queryset.model._meta.fields]
            writer.writerow(fields)
        
        # Write data
        for obj in queryset:
            row = [getattr(obj, field) for field in fields]
            writer.writerow(row)
        
        return response
    
    def export_excel(self, queryset, fields):
        """Export data as Excel."""
        import openpyxl
        from django.http import HttpResponse
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.basename}.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        # Write header
        if not fields:
            fields = [f.name for f in queryset.model._meta.fields]
        
        for col, field in enumerate(fields, 1):
            worksheet.cell(row=1, column=col, value=field)
        
        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col, field in enumerate(fields, 1):
                value = getattr(obj, field)
                worksheet.cell(row=row_num, column=col, value=str(value) if value else '')
        
        workbook.save(response)
        return response
    
    def export_pdf(self, queryset, fields):
        """Export data as PDF."""
        # Implementation would use reportlab
        return self.get_error_response("PDF export not implemented")


class FilterMixin:
    """
    Enhanced filtering mixin.
    Provides advanced filtering capabilities.
    """
    
    def get_queryset(self):
        """Apply custom filters from query params."""
        queryset = super().get_queryset()
        
        # Date range filtering
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        
        if date_from and hasattr(queryset.model, 'created_at'):
            queryset = queryset.filter(created_at__gte=date_from)
        
        if date_to and hasattr(queryset.model, 'created_at'):
            queryset = queryset.filter(created_at__lte=date_to)
        
        # Status filtering
        status = self.request.query_params.get('status')
        if status and hasattr(queryset.model, 'status'):
            queryset = queryset.filter(status=status)
        
        # Active/Inactive filtering
        is_active = self.request.query_params.get('is_active')
        if is_active is not None and hasattr(queryset.model, 'is_active'):
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset


class PermissionMixin:
    """
    Enhanced permission checking mixin.
    Provides role-based access control.
    """
    
    def check_object_permissions(self, request, obj):
        """Check object-level permissions."""
        super().check_object_permissions(request, obj)
        
        # Check ownership
        if hasattr(obj, 'created_by') and obj.created_by != request.user:
            # Check if user has special permission
            if not request.user.has_perm(f'{obj._meta.app_label}.change_{obj._meta.model_name}'):
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("You don't have permission to access this object")


class ReadOnlyViewSet(mixins.RetrieveModelMixin,
                     mixins.ListModelMixin,
                     viewsets.GenericViewSet):
    """
    ViewSet for read-only operations.
    Only allows GET requests.
    """
    pass
